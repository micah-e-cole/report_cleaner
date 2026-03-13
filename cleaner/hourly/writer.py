import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


import re

import re

import re

def split_room_fields(building_name: str, room_full: str):
    """
    Robust EMS room parser for all buildings including:
      - Fine Arts Building (and FINR abbreviation)
      - Administration Building
      - Pigott Building
      - 1103 E. Madison Building (special-case normalization)
      - All 'Other Area #x' rooms
      - Floor-level and named rooms with no dash

    Guarantees:
      - Building name NEVER appears in Room Number or Classroom Type.
      - Room Number always contains the meaningful identifier.
      - Classroom Type defaults to 'Other' when blank.
    """

    # ----------------------------------------------------------------------
    # 1. Normalize building name (strip suffix)
    # ----------------------------------------------------------------------
    room_full = str(room_full).strip()

    m_suffix = re.search(r"\s*\(Not REG Scheduled Space\)\s*$", building_name)
    if m_suffix:
        suffix = m_suffix.group(0)
        base_building = building_name[: m_suffix.start()]
    else:
        base_building = building_name
        suffix = ""

    # ----------------------------------------------------------------------
    # 2. Special-case for 1103 E. Madison Building
    # ----------------------------------------------------------------------
    display_building = base_building
    room_prefix_to_strip = None

    if base_building.startswith("1103 E. Madison"):
        m_madison = re.search(r"Madison(.*)", base_building)
        display_building = "Madison" + (m_madison.group(1) if m_madison else "")
        room_prefix_to_strip = "1103 E. Madison"

    # ----------------------------------------------------------------------
    # 3. FINR abbreviation → Fine Arts Building
    # ----------------------------------------------------------------------
    if room_full.startswith("FINR "):
        # Replace building context:
        display_building = "Fine Arts Building" + suffix
        # Remove FINR prefix so room parsing works:
        room_full = room_full.replace("FINR", "Fine Arts", 1)

    # Add suffix back
    display_building = display_building + suffix

    # ----------------------------------------------------------------------
    # 4. Strip building words from start of room_full
    # ----------------------------------------------------------------------
    room = room_full

    # 4a. For the Madison building special-case
    if room_prefix_to_strip and room.startswith(room_prefix_to_strip):
        room = room[len(room_prefix_to_strip):].lstrip()

    # 4b. Strip each word in the building name IF at start of the room string
    # This handles:
    #   Fine Arts 114 - ...
    #   Fine Arts 1st Floor Lobby
    #   Administration Other Area #1
    base_tokens = base_building.split()
    for tok in base_tokens:
        if room.startswith(tok + " "):
            room = room[len(tok):].lstrip()

    # Also strip uppercase variants ('FINE', 'ARTS', etc.)
    for tok in base_tokens:
        up_tok = tok.upper()
        if room.upper().startswith(up_tok + " "):
            room = room[len(tok):].lstrip()

    # ----------------------------------------------------------------------
    # 5. Handle "Other Area #x"
    # ----------------------------------------------------------------------
    if "Other Area" in room and " - " not in room:
        return display_building, room.strip(), "Other"

    # ----------------------------------------------------------------------
    # 6. Handle dash-separated rooms: "<number> - <type>"
    # ----------------------------------------------------------------------
    if " - " in room:
        left, right = room.split(" - ", 1)
        left = left.strip()
        right = right.strip()

        # Extract number-like token from left
        m_num = re.search(r"\b(\d+[A-Za-z0-9/-]*)\b", left)
        room_number = m_num.group(1) if m_num else left
        room_type = right or "Other"
        return display_building, room_number.strip(), room_type.strip()

    # ----------------------------------------------------------------------
    # 7. Handle floor-level and named rooms with no dash:
    #     1st Floor Lobby
    #     2nd Floor Lobby
    #     COSTUME SHOP
    # ----------------------------------------------------------------------
    room_number = room
    room_type = ""

    # If possible, identify number-like tokens (1st, 2nd, 201, etc.)
    # but keep the rest as type or default to Other
    # Example: "1st Floor Lobby" → room_number = "1st Floor Lobby"
    # Example: "COSTUME SHOP" → room_number = "COSTUME SHOP"
    # Example: "209 - Office" already handled above

    if not room_type.strip():
        room_type = "Other"

    return display_building, room_number.strip(), room_type.strip()

def format_hour_label(label: str) -> str:
    """Convert 'xa' -> 'x AM', 'xp' -> 'x PM'."""
    m = re.match(r"^(\d+)([ap])$", str(label).strip().lower())
    if not m:
        return label
    h = int(m.group(1))
    ap = m.group(2)
    return f"{h} {'AM' if ap == 'a' else 'PM'}"


def write_hourly_excel(df_long: pd.DataFrame, output_path: str) -> None:
    """
    Take the long-format hourly DataFrame from transform_hourly_utilization
    and write it to an Excel sheet with:

      Row 1: Title 'Seattle University, Reporting Period: ...'
      Row 2: Note 'All figures are percentages (Columns A:U).'
      Row 3: Headers
      Row 4+: Data

    Columns:
      A: Building
      B: Room Number
      C: Classroom Type
      D:...: Hour columns (e.g., 6 AM:9 PM depending on data)
      Last: Average Utilization by Room (row-wise average across hour columns)

    IMPORTANT: This writer is designed so that rooms with NO hourly data
               (all hour cells empty/NaN in EMS) STILL APPEAR in the output
               with 0.0 across all hour columns and Average.
    """
    # 1. Extract reporting period once; remove from data table
    reporting_period = None
    if "Reporting Period" in df_long.columns and not df_long["Reporting Period"].isna().all():
        reporting_period = df_long["Reporting Period"].dropna().iloc[0]
        df_long = df_long.drop(columns=["Reporting Period"])

    # 2. Derive Building, Room Number, Classroom Type
    df = df_long.copy()
    buildings = []
    room_numbers = []
    room_types = []

    for _, row in df.iterrows():
        b, rn, rt = split_room_fields(row["Building"], row["Room"])
        buildings.append(b)
        room_numbers.append(rn)
        room_types.append(rt)

    df["Building"] = buildings
    df["Room Number"] = room_numbers
    df["Classroom Type"] = room_types

    df["Classroom Type"] = (
        df["Classroom Type"]    # if there is no value, return "Other"
        .replace("", pd.NA)
        .fillna("Other")
    )

    df = df.drop(columns=["Room"])

    # Build a base list of all unique rooms (even if they have no hour data)
    base_index = df[["Building", "Room Number", "Classroom Type"]].drop_duplicates()

    # 3. Pivot to wide format: one row per room, columns per Hour
    if df.empty:
        wide = pd.DataFrame(columns=["Building", "Room Number", "Classroom Type"])
    else:
        # Keep raw Value entries, do NOT scale for percentages
        wide_pivot = df.pivot_table(
            index=["Building", "Room Number", "Classroom Type"],
            columns="Hour",
            values="Value",
            aggfunc=lambda x: x.iloc[0],  # keep first occurrence as-is
        )
        wide_pivot = wide_pivot.reset_index()

        # Flatten possible MultiIndex columns from pivot
        wide_pivot.columns = [
            col if isinstance(col, str) else col[1] for col in wide_pivot.columns
        ]

        # Merge base index with pivot to ensure rooms with no data are retained
        wide = base_index.merge(
            wide_pivot,
            on=["Building", "Room Number", "Classroom Type"],
            how="left",
        )

        # Drop any 'Average' column coming from EMS (pivot),
        # we'll compute our own Average later.
        if "Average" in wide.columns:
            wide = wide.drop(columns=["Average"])

        # Force expected hour columns to exist even if empty
        # We support both 6a–8p and 7a–9p style ranges by including a superset.
        expected_hours = [
            "6a", "7a", "8a", "9a", "10a", "11a",
            "12p", "1p", "2p", "3p", "4p", "5p",
            "6p", "7p", "8p", "9p",
        ]
        for h in expected_hours:
            if h not in wide.columns:
                wide[h] = pd.NA

        # Sort hour columns logically
        def hour_key(label):
            if label in ("Building", "Room Number", "Classroom Type"):
                return -1
            m = re.match(r"^(\d+)([ap])$", str(label).strip().lower())
            if not m:
                return 999
            hh = int(m.group(1))
            ap = m.group(2)
            if ap == "a":
                h24 = hh if hh != 12 else 0
            else:
                h24 = hh + 12 if hh != 12 else 12
            return h24

        base_cols = ["Building", "Room Number", "Classroom Type"]
        hour_cols = [c for c in wide.columns if c not in base_cols]
        hour_cols_sorted = sorted(hour_cols, key=hour_key)

        # Ensure hour columns are numeric; empty/NaN → 0
        for h in hour_cols_sorted:
            wide[h] = pd.to_numeric(wide[h], errors="coerce").fillna(0)

        # Add Average column (will be the ONLY 'Average' column now)
        wide["Average"] = 0.0

        # Reorder columns: A–C base, D–? hours, last = Average
        wide = wide[base_cols + hour_cols_sorted + ["Average"]]

        # Rename hour labels for display (6a → 6 AM, etc.)
        rename_map = {h: format_hour_label(h) for h in hour_cols_sorted}
        wide = wide.rename(columns=rename_map)
        # Average column header stays "Average"

        # OPTIONAL: if you want to append "(Not REG Scheduled Space)" here
        # instead of (or in addition to) transform-level tagging:
        #
        # Find columns that are hours (after renaming)
        hour_display_cols = [format_hour_label(h) for h in expected_hours]
        hour_display_cols = [c for c in hour_display_cols if c in wide.columns]

        # A room is "Not REG Scheduled Space" if ALL hour values are 0.0
        no_data_mask = (wide[hour_display_cols] == 0).all(axis=1)

        wide.loc[no_data_mask, "Building"] = (
            wide.loc[no_data_mask, "Building"] + " (Not REG Scheduled Space)"
        )

    sheet_name = "Hourly Utilization"

    # 4. Write wide DataFrame starting at row 3 (startrow=2 => header at row3)
    wide.to_excel(output_path, index=False, sheet_name=sheet_name, startrow=2)

    # 5. Apply Excel formatting and formulas
    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    max_col = ws.max_column
    end_col = max_col

    # Title row (row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = (
        f"Seattle University, Reporting Period: {reporting_period}"
        if reporting_period
        else "Seattle University, Hourly Room Utilization"
    )
    ws["A1"].font = Font(bold=True)

    # Note row (row 2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = "All figures are percentages (Columns A:U)."

    # Header row (row 3) bold
    for cell in ws[3]:
        cell.font = Font(bold=True)

    # Freeze panes below header
    ws.freeze_panes = "A4"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Auto-ish column widths based on content in rows 3+
    for col_idx, col_cells in enumerate(
        ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Set up Average in last column (row-wise, across hour columns)
    first_hour_col_idx = 4   # D = first hour column
    last_hour_col_idx = max_col - 1  # last hour col (before Average)
    avg_col_idx = max_col          # Average column

    first_hour_letter = get_column_letter(first_hour_col_idx)
    last_hour_letter = get_column_letter(last_hour_col_idx)
    avg_letter = get_column_letter(avg_col_idx)

    # Data rows start at row 4 (since row 3 is the header)
    for row_idx in range(4, ws.max_row + 1):
        ws[f"{avg_letter}{row_idx}"] = (
            f"=AVERAGE({first_hour_letter}{row_idx}:{last_hour_letter}{row_idx})"
        )

    # Format hour and Average columns as standard numeric (no percentage scaling)
    for col_idx in range(first_hour_col_idx, avg_col_idx + 1):
        for cell in ws.iter_rows(
            min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            cell[0].number_format = "0.0"

    wb.save(output_path)