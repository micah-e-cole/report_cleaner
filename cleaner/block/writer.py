from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def write_formatted_excel(df, output_path: str):
    sheet_name = "Classroom Utilization"

    df.to_excel(output_path, index=False, sheet_name=sheet_name)

    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    # Bold header
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Auto column widths
    for col_idx, col_cells in enumerate(ws.iter_cols(
        min_row=1, max_row=ws.max_row,
        min_col=1, max_col=ws.max_column
    ), start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Percent columns
    percent_cols = ['Utilization %', 'Seat Fill %']
    for col_name in percent_cols:
        if col_name in df.columns:
            col_idx = list(df.columns).index(col_name) + 1
            for cell in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                     min_col=col_idx, max_col=col_idx):
                cell[0].number_format = '0.0%'

    # Timestamp
    ws['J1'] = f"Date: {datetime.today().strftime('%m/%d/%Y %I:%M %p')}"

    wb.save(output_path)
