# excel_cleaner/cleaner/logic.py
# ---------------- ABOUT  ----------------
# Author: Micah Braun
# AI Acknowledgement: This file was compiled with assistance from
#                     Copilot alongside Enterprise Data Protection.
# Date: 02/24/2026

# ---------------- LIBRARIES ----------------
import os
import pandas as pd
import xlrd
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

ALLOWED_EXTENSIONS = {'.csv', '.xls', '.xlsx'}


# ---------------- FILE HELPER FUNCTIONS ----------------
def is_valid_input_file(path: Path) -> bool:
    """
    Return True if the file exists, is a file (not a directory),
    and has an allowed extension.
    """
    return path.is_file() and path.suffix.lower() in ALLOWED_EXTENSIONS


def collect_input_files(paths: list[str]) -> list[Path]:
    """
    Given a list of path strings (files or directories),
    return a list of valid files to process.

    - If path is a directory: include all allowed files in that directory
    - If path is a file: include it if its extension is allowed
    - If extension is not allowed: print a warning and skip
    """
    collected: list[Path] = []

    for p_str in paths:
        p = Path(p_str)

        if not p.exists():
            print(f"⚠️  Skipping '{p}': path does not exist.")
            continue

        if p.is_dir():
            # Process all allowed files in the directory
            for file in p.iterdir():
                if is_valid_input_file(file):
                    collected.append(file)
                elif file.is_file():
                    print(
                        f"⚠️  Skipping '{file.name}': "
                        f"'.{file.suffix.lstrip('.')}' is not an accepted file type. "
                        f"Accepted types are: {', '.join(ALLOWED_EXTENSIONS)}"
                    )
        else:
            # It's a file
            if is_valid_input_file(p):
                collected.append(p)
            else:
                print(
                    f"⚠️  Skipping '{p.name}': "
                    f"'.{p.suffix.lstrip('.')}' is not an accepted file type. "
                    f"Accepted types are: {', '.join(ALLOWED_EXTENSIONS)}"
                )

    return collected


# ---------------- CORE TRANSFORM FUNCTIONS ----------------
def transform_classroom_utilization(input_path: str) -> pd.DataFrame:
    """
    Transform an EMS classroom utilization .XLSX/.XLS/.CSV export into a normalized DataFrame.

    This function:
      - Removes repeated header noise and timestamp rows.
      - Identifies building header rows and forward-fills building names.
      - Extracts room-level data rows.
      - Converts numeric and percentage fields into consistent formats.

    Args:
        input_path (str):
            Path to the raw EMS file containing the exported utilization report.

    Returns:
        pandas.DataFrame:
            A cleaned table with one row per room containing:
            ['Building', 'Room', 'Class Meetings', 'Class Hours',
             'Utilization %', 'Avg Est Enroll', 'Avg Act Enroll',
             'Max Capacity', 'Seat Fill %']

    Raises:
        FileNotFoundError:
            If the given file path does not exist.
        ValueError:
            If the input file contents cannot be parsed into the expected structure.
    """

    ext = os.path.splitext(input_path)[1].lower()

    # 1. Choose correct reader for file type
    if ext == ".csv":
        df = pd.read_csv(input_path, header=None)
    elif ext == ".xlsx":
        df = pd.read_excel(input_path, header=None, engine="openpyxl")
    elif ext == ".xls":
        df = pd.read_excel(input_path, header=None, engine="xlrd")
    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .xls, .xlsx, or .csv.")

    # 2. Drop fully empty rows
    df2 = df.dropna(how='all').copy()

    # 3. Build row strings to detect header / noise rows
    def row_to_str(row):
        return ' '.join([str(x) for x in row if pd.notna(x)])

    row_str = df2.apply(row_to_str, axis=1)

    noise_keywords = [
        'Seattle University',
        'Reporting Period',
        'Classroom Utilization',
        'Class Meetings',
        'Avg. Est.  Enroll',
        'Avg. Est. Enroll',
        'Avg. Act.  Enroll',
        'Avg. Act. Enroll',
        'Seat Fill',
        'Page ',
    ]
    base_noise = row_str.str.contains('|'.join(noise_keywords))

    # Timestamp rows like "2/19/2026 3:39 PM RZ"
    mask_date = row_str.str.contains(r'\d{1,2}/\d{1,2}/\d{4} \d{1,2}:\d{2} [AP]M')

    # Keep everything that is not noise or timestamp
    clean = df2[~(base_noise | mask_date)].copy()

    # 4. Attach Building name by forward-filling building rows
    clean['Building'] = None
    current_building = None

    for idx, row in clean.iterrows():
        bname = row[0]
        room = row[1]

        # A building row: text in col0, col1 is NaN, and not a Total/Average line
        if pd.notna(bname) and pd.isna(room) and not str(bname).startswith(('Total for', 'Average for')):
            current_building = bname
            clean.at[idx, 'Building'] = current_building
        else:
            clean.at[idx, 'Building'] = current_building

    # 5. Room-level rows: have room name (col1) and numeric values in cols 4 and 6
    room_rows = clean[clean[1].notna() & clean[4].notna() & clean[6].notna()].copy()

    # 6. Build clean table
    out = pd.DataFrame({
        'Building':       room_rows['Building'].values,
        'Room':           room_rows[1].values,
        'Class Meetings': room_rows[4].values,
        'Class Hours':    room_rows[6].values,
        'Utilization %':  room_rows[7].values,
        'Avg Est Enroll': room_rows[8].values,
        'Avg Act Enroll': room_rows[9].values,
        'Max Capacity':   room_rows[11].values,
        'Seat Fill %':    room_rows[12].values,
    })

    # 7. Convert numeric columns
    for col in ['Class Meetings', 'Class Hours', 'Avg Est Enroll', 'Avg Act Enroll', 'Max Capacity']:
        out[col] = pd.to_numeric(out[col], errors='coerce')

    # 8. Convert values like "97.33%" -> 0.9733
    for col in ['Utilization %', 'Seat Fill %']:
        # Treat incoming values as strings to detect literal percent signs
        col_str = out[col].astype(str)

        # Identify rows like "97.33%"
        mask_pct = col_str.str.contains('%', na=False)

        # Create a numeric Series aligned with the DataFrame index
        numeric = pd.Series(index=out.index, dtype='float64')

        # Percent-case: "97.33%" -> 0.9733
        numeric[mask_pct] = (
            col_str[mask_pct]
            .str.replace('%', '', regex=False)
            .astype(float) / 100.0
        )

        # Non-percent-case: treat as numeric directly
        numeric[~mask_pct] = pd.to_numeric(out.loc[~mask_pct, col], errors='coerce')

        # If some values are like 97.33 instead of 0.9733, fix them
        numeric = numeric.where((numeric <= 1) | numeric.isna(), numeric / 100.0)

        # Assign the clean numeric series back to the column
        out[col] = numeric

    return out


def write_formatted_excel(df: pd.DataFrame, output_path: str):
    """
    Write a cleaned and normalized classroom utilization DataFrame to Excel.

    This function:
      - Writes headers beginning at cell A1.
      - Applies bold formatting to headers.
      - Freezes the header row for easier navigation.
      - Applies percentage formatting to relevant columns.
      - Auto-adjusts column widths.
      - Inserts the run date into cell J1.

    Args:
        df (pandas.DataFrame):
            The cleaned utilization data returned from transform_classroom_utilization().
        output_path (str):
            File path where the Excel file will be written.

    Returns:
        None
    """
    # 1. Write basic Excel (pandas puts headers at row 1, starting A1)
    sheet_name = "Classroom Utilization"
    df.to_excel(output_path, index=False, sheet_name=sheet_name)

    # 2. Load with openpyxl to apply formatting
    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    # Bold header row (row 1)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Freeze panes below header (so row 1 stays visible)
    ws.freeze_panes = "A2"

    # Add autofilter over entire data range
    ws.auto_filter.ref = ws.dimensions

    # Auto-ish column widths based on content
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1,
                                                     max_row=ws.max_row,
                                                     min_col=1,
                                                     max_col=ws.max_column), start=1):
        max_len = 0
        for cell in col_cells:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Apply percentage number format to Utilization % and Seat Fill % columns
    percent_cols = ['Utilization %', 'Seat Fill %']
    for col_name in percent_cols:
        if col_name in df.columns:
            col_idx = list(df.columns).index(col_name) + 1  # 1-based index
            for cell in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                     min_col=col_idx, max_col=col_idx):
                cell[0].number_format = '0.0%'

    # 3. Set J1 = "Date: MM/DD/YYYY HH:MM AM/PM"
    today_str = datetime.today().strftime('%m/%d/%Y %I:%M %p')
    ws['J1'] = f"Date: {today_str}"

    wb.save(output_path)


def run_cleaner(input_xlsx: str, output_xlsx: str) -> None:
    """
    Process a raw EMS classroom utilization export and write a cleaned Excel report.

    Args:
        input_xlsx (str): Path to the raw exported input file from EMS.
        output_xlsx (str): Destination path where the formatted Excel file will be written.

    Returns:
        None: The function writes an Excel file but returns nothing.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If the input file structure does not match expected format.
    """
    df = transform_classroom_utilization(input_xlsx)
    write_formatted_excel(df, output_xlsx)


def run_batch_cleaner(inputs: list[str], output_dir: str | None = None) -> None:
    """
    Batch-process one or more input paths (files and/or directories).

    - Only .csv, .xls, .xlsx files will be processed.
    - Any other file types will trigger a warning and be skipped.
    - Output files are written as '<original_stem>_cleaned.xlsx'.

    Args:
        inputs (list[str]):
            List of file and/or directory paths.
        output_dir (str | None):
            Optional directory where all cleaned files will be written.
            If None, each file is written next to its input.
    """
    files_to_process = collect_input_files(inputs)

    if not files_to_process:
        print("No valid input files found. Exiting.")
        return

    if output_dir is not None:
        out_base = Path(output_dir)
        out_base.mkdir(parents=True, exist_ok=True)
    else:
        out_base = None  # meaning: use each file's parent

    print(f"Found {len(files_to_process)} file(s) to process.")

    for f in files_to_process:
        try:
            if out_base is None:
                target_dir = f.parent
            else:
                target_dir = out_base

            output_path = target_dir / f"{f.stem}_cleaned.xlsx"

            print(f"🔄 Processing: {f}")
            run_cleaner(str(f), str(output_path))
            print(f"✅ Finished: {f.name} → {output_path}")
        except Exception as e:
            # Don't crash entire batch for one bad file
            print(f"❌ Error processing '{f}': {e}")

    print("🎉 Batch processing complete.")


# Command-line entry point
if __name__ == "__main__":
    # Command-line usage:
    #   python logic.py <file_or_folder> [more_files_or_folders...] [--out OUTPUT_DIR]
    #
    # Examples:
    #   python logic.py "C:/path/to/one_file.xlsx"
    #   python logic.py "C:/path/to/folder_with_exports"
    #   python logic.py "file1.csv" "file2.xlsx" --out "C:/output/cleaned"
    #
    args = sys.argv[1:]

    if not args:
        print("Usage: python logic.py <file_or_folder> [more_files_or_folders...] [--out OUTPUT_DIR]")
        print("Accepted file types:", ", ".join(ALLOWED_EXTENSIONS))
        sys.exit(1)

    # Simple parsing for optional --out argument
    if "--out" in args:
        out_index = args.index("--out")
        input_args = args[:out_index]
        if out_index + 1 >= len(args):
            print("Error: --out provided but no output directory specified.")
            sys.exit(1)
        output_dir = args[out_index + 1]
    else:
        input_args = args
        output_dir = None

    run_batch_cleaner(input_args, output_dir)