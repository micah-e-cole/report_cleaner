# tests/test_hourly_writer.py

import pandas as pd
from openpyxl import load_workbook
from cleaner.hourly.writer import write_hourly_excel


def make_sample_long_df():
    return pd.DataFrame(
        {
            "Building": ["Bannan", "Bannan", "Bannan"],
            "Room": [
                "Bannan 222 - Classroom",
                "Bannan 222 - Classroom",
                "Bannan 222 - Classroom",
            ],
            "Hour": ["6a", "7a", "Average"],
            "Value": [1.0, 2.0, 10.0],
            "Reporting Period": ["Fall 2025", "Fall 2025", "Fall 2025"],
        }
    )


def test_write_hourly_excel_single_average_column(tmp_path):
    df_long = make_sample_long_df()
    out_path = tmp_path / "hourly_out.xlsx"

    write_hourly_excel(df_long, out_path)

    wb = load_workbook(out_path)
    ws = wb["Hourly Utilization"]

    # Header row is row 3
    headers = [cell.value for cell in ws[3]]
    # Expect: Building, Room Number, Classroom Type, 6 AM, 7 AM, ..., Average
    assert "Average" in headers
    assert headers.count("Average") == 1

    # Find Average column index
    avg_col_idx = headers.index("Average") + 1  # 1-based
    avg_col_letter = ws.cell(row=3, column=avg_col_idx).column_letter

    # Data row is row 4
    avg_formula = ws[f"{avg_col_letter}4"].value
    # Should be something like =AVERAGE(D4:R4)
    assert isinstance(avg_formula, str)
    assert avg_formula.startswith("=AVERAGE(")
    assert ":".join(part for part in avg_formula.split("AVERAGE(")[1].rstrip(")").split(":")).endswith("4")

    # The numeric values that were written must be present in hour cells
    # We don't assert exact positions here, just that some hour columns have 1.0 and 2.0.
    row4_values = [cell.value for cell in ws[4]]
    assert 1.0 in row4_values
    assert 2.0 in row4_values