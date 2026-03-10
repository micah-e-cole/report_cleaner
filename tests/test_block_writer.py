# tests/test_block_writer.py

import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pandas as pd
from openpyxl import load_workbook

from cleaner.block.writer import write_formatted_excel  # or your actual writer name


def test_block_writer_creates_excel(tmp_path):
    df = pd.DataFrame({
        "Building": ["A"],
        "Room": ["101"],
        "Hour": ["8a"],
        "Value": [2.0],
    })

    out_file = tmp_path / "block_out.xlsx"
    write_formatted_excel(df, out_file)

    wb = load_workbook(out_file)
    # Match real behavior: sheet is "Classroom Utilization"
    assert "Classroom Utilization" in wb.sheetnames