# tests/test_logic.py

import os
from pathlib import Path

import pandas as pd
import openpyxl
from cleaner.logic import (
transform_classroom_utilization,
    write_formatted_excel,
    run_cleaner,
    run_batch_cleaner,
    collect_input_files,
    ALLOWED_EXTENSIONS,
)


def create_sample_csv(path: Path):
    """Create a minimal CSV mimicking EMS export structure."""
    # Columns by index as expected in transform_classroom_utilization
    # We'll create a small DataFrame with:
    # - one header/noise row
    # - one building row
    # - one room row
    rows = [
        ["Seattle University", None, None, None, None, None, None, None, None, None, None, None, None],
        ["Pigott Building", None, None, None, None, None, None, None, None, None, None, None, None],
        [None, "PIGT 103", None, None, 10, None, 20, "97.33%", 25, 24, None, 30, "80%"],
    ]
    df = pd.DataFrame(rows)
    df.to_csv(path, index=False, header=False)


def test_transform_classroom_utilization_basic(tmp_path):
    input_file = tmp_path / "sample.csv"
    create_sample_csv(input_file)

    df = transform_classroom_utilization(str(input_file))

    # Expected columns present
    expected_cols = [
        "Building",
        "Room",
        "Class Meetings",
        "Class Hours",
        "Utilization %",
        "Avg Est Enroll",
        "Avg Act Enroll",
        "Max Capacity",
        "Seat Fill %",
    ]
    assert list(df.columns) == expected_cols

    # One row of data
    assert len(df) == 1

    row = df.iloc[0]
    assert row["Building"] == "Pigott Building"
    assert row["Room"] == "PIGT 103"
    assert row["Class Meetings"] == 10
    assert row["Class Hours"] == 20
    # 97.33% should be 0.9733
    assert round(row["Utilization %"], 4) == 0.9733
    # 80% should be 0.8
    assert row["Seat Fill %"] == 0.8


def test_transform_rejects_unsupported_extension(tmp_path):
    bad_file = tmp_path / "sample.txt"
    bad_file.write_text("just some text")

    try:
        transform_classroom_utilization(str(bad_file))
        assert False, "Expected ValueError for unsupported extension"
    except ValueError as e:
        assert ".txt" in str(e)
        for ext in [".xls", ".xlsx", ".csv"]:
            assert ext in str(e)


def test_write_formatted_excel_creates_file_and_sheet(tmp_path):
    # Create a simple DataFrame with required columns
    df = pd.DataFrame(
        {
            "Building": ["Building A"],
            "Room": ["101"],
            "Class Meetings": [5],
            "Class Hours": [10],
            "Utilization %": [0.5],
            "Avg Est Enroll": [20],
            "Avg Act Enroll": [18],
            "Max Capacity": [30],
            "Seat Fill %": [0.6],
        }
    )

    output_file = tmp_path / "output.xlsx"
    write_formatted_excel(df, str(output_file))

    assert output_file.exists()

    # Load with openpyxl to inspect
    wb = openpyxl.load_workbook(output_file)
    assert "Classroom Utilization" in wb.sheetnames
    ws = wb["Classroom Utilization"]

# Headers in the first len(columns) cells of row 1
    header_cells = list(ws[1])[: len(df.columns)]
    headers = [cell.value for cell in header_cells]
    assert headers == list(df.columns)

    # J1 contains the Date label (could be an extra cell beyond the headers)
    assert ws["J1"].value.startswith("Date:")


def test_collect_input_files_only_accepts_allowed(tmp_path, capsys):
    # Create some files
    valid_csv = tmp_path / "a.csv"
    valid_xlsx = tmp_path / "b.xlsx"
    invalid_pdf = tmp_path / "c.pdf"

    for p in [valid_csv, valid_xlsx, invalid_pdf]:
        p.write_text("dummy")

    from excel_cleaner.cleaner.logic import collect_input_files

    files = collect_input_files([str(tmp_path)])
    # Should include only allowed extensions
    names = {f.name for f in files}
    assert "a.csv" in names
    assert "b.xlsx" in names
    assert "c.pdf" not in names

    # Optionally verify that a warning was printed for the pdf
    captured = capsys.readouterr()
    assert ".pdf" in captured.out
    for ext in ALLOWED_EXTENSIONS:
        assert ext in captured.out


def test_run_cleaner_creates_output_file(tmp_path):
    input_file = tmp_path / "sample.csv"
    create_sample_csv(input_file)

    output_file = tmp_path / "sample_Clean.xlsx"
    run_cleaner(str(input_file), str(output_file))

    assert output_file.exists()


def test_run_batch_cleaner_creates_cleaned_files(tmp_path):
    # Create two sample input files
    f1 = tmp_path / "sample1.csv"
    f2 = tmp_path / "sample2.csv"
    create_sample_csv(f1)
    create_sample_csv(f2)

    out_dir = tmp_path / "out"
    run_batch_cleaner([str(f1), str(f2)], str(out_dir))

    # Expect two files: sample1_cleaned.xlsx and sample2_cleaned.xlsx
    out_files = {p.name for p in out_dir.iterdir()}
    assert "sample1_cleaned.xlsx" in out_files
    assert "sample2_cleaned.xlsx" in out_files